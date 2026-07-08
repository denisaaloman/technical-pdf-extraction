import csv
import io
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def build_wide_rows(results: List[Dict[str, Any]]) -> Tuple[List[str], List[Dict[str, Any]]]:
    columns: List[str] = []
    seen = set()
    rows: List[Dict[str, Any]] = []

    for doc in results:
        for table in doc.get("tables", []):
            for header in table.get("headers", []):
                if header not in seen:
                    seen.add(header)
                    columns.append(header)
            for row in table.get("rows", []):
                rows.append(
                    {
                        "document": doc.get("filename", ""),
                        "tableTitle": table.get("title") or "(tabel fără titlu)",
                        "values": row,
                    }
                )

    return columns, rows


def to_csv(columns: List[str], rows: List[Dict[str, Any]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Document", "Categorie", *columns])
    for r in rows:
        writer.writerow(
            [r["document"], r["tableTitle"], *[r["values"].get(c, "") for c in columns]]
        )
    return output.getvalue()


def to_xlsx(columns: List[str], rows: List[Dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Technical Tables"

    header_row = ["Document", "Categorie", *columns]
    sheet.append(header_row)
    for cell in sheet[1]:
        cell.font = cell.font.copy(bold=True)

    for r in rows:
        sheet.append([r["document"], r["tableTitle"], *[r["values"].get(c, "") for c in columns]])

    sheet.freeze_panes = "C2"

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 26
    for i, _ in enumerate(columns, start=3):
        sheet.column_dimensions[get_column_letter(i)].width = 20

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
